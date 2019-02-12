import json
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def dct_from_excel(path, sheet_name):
    wb = load_workbook(path)
    current_sheet = wb[sheet_name]
    result_dict = []
    for i in range(1, current_sheet.max_row):
        line = dict()
        x = 0
        for header in list(current_sheet.iter_rows())[0]:
            val = list(current_sheet.iter_rows())[i][x].value
            if val:
                line[header.value] = val
            x += 1
        result_dict.append(line)
    return result_dict


def get_ip(filename,
           user,
           password,
           host,
           port,
           token,
           ):
    try:
        physical_dct = dct_from_excel('fssp.xlsx', "physical")
        legal_dct = dct_from_excel('fssp.xlsx', "legal")
        ip_dct = dct_from_excel('fssp.xlsx', "ip")
        if not (physical_dct or legal_dct or ip_dct):
            print(f'Не найдены данные для запроса в файле {filename}')
            return None
    except Exception as err:
        print(f'Проверьте входной файл - {filename}. Ошибка - {err}')
        return None
    if user:
        proxies = dict(https='https://{user}:{password}@{host}:{port}'.format(user=user,
                                                                              password=password,
                                                                              host=host,
                                                                              port=port,
                                                                              )
                       )
    else:
        proxies = None

    headers = {"Accept": "application/json",
               "Content-Type": "application/json",
               "charset": "utf-8",
               }

    base_url = 'https://api-ip.fssprus.ru/api/v1.0'

    requests_data = []
    for item in physical_dct:
        requests_data.append(dict(type=1, params=item))
    for item in legal_dct:
        requests_data.append(dict(type=2, params=item))
    for item in ip_dct:
        requests_data.append(dict(type=3, params=item))

    data_to_post = dict(token=token, request=requests_data)
    data_to_post = json.dumps(data_to_post).encode("utf-8")

    request = requests.post(url=f'{base_url}/search/group',
                            headers=headers,
                            data=data_to_post,
                            proxies=proxies,
                            )

    if request.status_code == 200:
        response = request.json().get('response')
        task = response.get('task')
        progress = None
        print(f'Ожидаем выполнения запроса task - {task}')

        is_ready_status = False
        while not is_ready_status:
            request = requests.get(url=f'{base_url}/status?'
                                       f'token={token}&'
                                       f'task={task}',
                                   proxies=proxies,
                                   headers=headers)
            response = request.json().get('response')

            if progress is None or response and progress != response.get("progress"):
                progress = response.get("progress")
                print(f'Прогресс {progress}')
            if response and response.get('status') == 0:
                is_ready_status = True

        request = requests.get(url=f'{base_url}/result?'
                                   f'token={token}&'
                                   f'task={task}',
                               proxies=proxies,
                               headers=headers)

        if request.status_code == 200:

            response = request.json().get('response')
            result = response.get('result')

            wb = Workbook()

            filepath = "fssp_out.xlsx"
            sheet = wb.active
            tab_headers = ['name', 'exe_production', 'details', 'subject', 'department', 'bailiff', 'ip_end']
            sheet.append(tab_headers)
            for items in result:
                for item in items.get('result'):
                    sheet.append(list(item.values()))
            style = TableStyleInfo(name="TableStyleLight8",
                                   showFirstColumn=False,
                                   showLastColumn=False,
                                   showRowStripes=True,
                                   showColumnStripes=True)
            tab = Table(displayName="TableResult",
                        ref=f"A1:G{len(list(sheet))}",
                        tableStyleInfo=style,
                        )
            sheet.add_table(tab)
            try:
                wb.save(filepath)
                print(f'Сформирован файл {filepath}')
            except PermissionError:
                print(f'Ошибка записи. Закройте файл {filepath}')

        else:
            print({400: request.json().get('exception'),
                   401: request.json().get('exception')
                   }.get(request.status_code))
    else:
        print({400: request.json().get('exception'),
               401: request.json().get('exception')
               }.get(request.status_code))


if __name__ == '__main__':
    settings_dct = dict(token="TqMfEh4ffNsj",
                        host="host",
                        password="********",
                        port=3128,
                        user=None,
                        filename="fssp.xlsx",
                        )
    get_ip(**settings_dct)
    input("Press Enter to continue ...")
